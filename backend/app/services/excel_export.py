from __future__ import annotations
import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.user import UserRole


PIPELINE_STATUS_RU = {
    "active_work": "Активная работа",
    "on_visa": "На визе",
    "paused": "Пауза",
    "changed_mind": "Передумали",
    "refund": "На возврате",
    "unpaid": "Не оплачено",
    "transferred_pipeline": "Перевели",
    "ielts_retake": "Пересдача IELTS",
    "suspended": "Подвешено",
    "no_status": "Нет статуса",
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _header(ws, row: list[str]):
    ws.append(row)
    for cell in ws[ws.max_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def export_students_list(students: list[dict], role: UserRole) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Студенты"

    headers = ["Имя", "Телефон", "Степень", "Город", "Год поступления", "Статус", "Дней в работе"]
    _header(ws, headers)

    for s in students:
        ws.append([
            s.get("full_name"),
            s.get("phone"),
            s.get("degree_level"),
            s.get("city"),
            s.get("intake_year"),
            PIPELINE_STATUS_RU.get(s.get("pipeline_status", ""), s.get("pipeline_status", "")),
            s.get("days_in_work"),
        ])

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_student_card(student: dict, role: UserRole) -> bytes:
    wb = openpyxl.Workbook()

    # Profile sheet
    ws = wb.active
    ws.title = "Профиль"
    _header(ws, ["Поле", "Значение"])
    profile_fields = [
        ("ФИО", student.get("full_name")),
        ("Телефон", student.get("phone")),
        ("Город", student.get("city")),
        ("Возраст", student.get("age")),
        ("Степень", student.get("degree_level")),
        ("Специальность", student.get("specialty")),
        ("GPA", student.get("gpa")),
        ("Бюджет", student.get("budget_per_year")),
        ("Год поступления", student.get("intake_year")),
        ("Сезон", student.get("intake_season")),
        ("Достижения", student.get("achievements_text")),
    ]
    for k, v in profile_fields:
        ws.append([k, str(v) if v is not None else ""])
    _autofit(ws)

    # Applications sheet
    ws2 = wb.create_sheet("Подачи")
    _header(ws2, ["Страна", "Университет", "Запланировано", "Подано", "Статус подачи", "Статус визы", "Стипендия"])
    for app in student.get("applications", []):
        ws2.append([
            app.get("country"),
            app.get("university", ""),
            app.get("submissions_planned"),
            app.get("submissions_done"),
            app.get("submission_status"),
            app.get("visa_status", ""),
            "Да" if app.get("scholarship_target") else "Нет",
        ])
    _autofit(ws2)

    # Services sheet
    ws3 = wb.create_sheet("Услуги")
    _header(ws3, ["Тип", "Включена", "Статус", "Результат"])
    for svc in student.get("services", []):
        ws3.append([
            svc.get("service_type"),
            "Да" if svc.get("included") else "Нет",
            svc.get("status"),
            svc.get("result", ""),
        ])
    _autofit(ws3)

    # Finance sheet (admin/mzk only)
    if role in (UserRole.admin, UserRole.mzk_manager):
        for contract in student.get("contracts", []):
            ws4 = wb.create_sheet("Финансы")
            _header(ws4, ["Поле", "Значение"])
            finance_fields = [
                ("Сумма договора", contract.get("amount")),
                ("Дата подписания", contract.get("signed_date")),
                ("Статус", PIPELINE_STATUS_RU.get(contract.get("pipeline_status", ""), "")),
                ("Остаток", contract.get("client_remaining_amount")),
                ("Дата остатка", contract.get("client_remaining_date")),
                ("Итого менторам", contract.get("mentor_total_owed")),
                ("Сумма English", contract.get("english_sum")),
                ("Оплачено English", contract.get("english_paid")),
            ]
            for k, v in finance_fields:
                ws4.append([k, str(v) if v is not None else ""])

            ws5 = wb.create_sheet("Платежи")
            _header(ws5, ["Тип", "Сумма", "Статус", "Дата"])
            for p in contract.get("payments", []):
                ws5.append([
                    p.get("type"),
                    p.get("amount"),
                    p.get("status"),
                    p.get("paid_at", ""),
                ])
            _autofit(ws5)
            break  # Export first contract only

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
